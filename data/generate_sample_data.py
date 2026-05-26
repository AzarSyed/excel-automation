"""
Bootstrap script — generates realistic sample CSV files with intentional
data quality issues so the tool can demonstrate all cleaning/validation features.
Run once:  python data/generate_sample_data.py
"""
import random
import string
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd

SEED = 42
random.seed(SEED)
np.random.seed(SEED)

DATA_DIR = Path(__file__).parent


# ── Helpers ──────────────────────────────────────────────────────────────────

def rand_date(start: date, end: date) -> str:
    return (start + timedelta(days=random.randint(0, (end - start).days))).strftime("%Y-%m-%d")


def corrupt_email(email: str) -> str:
    """Return a deliberately malformed email string."""
    modes = [
        lambda e: e.replace("@", ""),
        lambda e: e.replace(".", ""),
        lambda e: e + "@extra@domain.com",
        lambda e: "not-an-email",
        lambda e: "  " + e,  # leading space (caught by trim)
    ]
    return random.choice(modes)(email)


# ── Sales dataset ─────────────────────────────────────────────────────────────

def generate_sales(n: int = 220) -> pd.DataFrame:
    products = {
        "Electronics": ["Laptop Pro 15", "Wireless Mouse", "USB-C Hub", "4K Monitor", "Mechanical Keyboard"],
        "Software": ["CRM License", "Analytics Suite", "Project Manager Pro", "Email Marketing Tool"],
        "Services": ["Onboarding Package", "Support Plan", "Custom Training", "Implementation"],
        "Office": ["Ergonomic Chair", "Standing Desk", "Label Printer", "Wireless Projector"],
    }
    regions = ["North", "South", "East", "West", "Central"]
    reps = ["Alice Chen", "Bob Martinez", "Carol Johnson", "David Kim", "Emma Wilson", "Frank Lee"]
    statuses = ["Completed", "Completed", "Completed", "Pending", "Cancelled", "Processing"]
    first_names = ["James", "Mary", "John", "Patricia", "Robert", "Jennifer", "Michael", "Linda",
                   "William", "Barbara", "David", "Susan", "Richard", "Jessica", "Joseph", "Sarah"]
    last_names = ["Smith", "Johnson", "Williams", "Brown", "Jones", "Garcia", "Miller", "Davis",
                  "Wilson", "Moore", "Taylor", "Anderson", "Thomas", "Jackson", "White", "Harris"]
    domains = ["gmail.com", "yahoo.com", "outlook.com", "company.com", "business.io", "corp.net"]

    rows = []
    for i in range(1, n + 1):
        cat = random.choice(list(products.keys()))
        product = random.choice(products[cat])
        qty = random.randint(1, 20)
        unit_price = round(random.uniform(9.99, 2499.99), 2)
        total = round(qty * unit_price, 2)
        fname = random.choice(first_names)
        lname = random.choice(last_names)
        email = f"{fname.lower()}.{lname.lower()}@{random.choice(domains)}"

        rows.append({
            "Order ID": f"ORD-{i:05d}",
            "Date": rand_date(date(2024, 1, 1), date(2024, 12, 31)),
            "Customer Name": f"{fname} {lname}",
            "Customer Email": email,
            "Product": product,
            "Category": cat,
            "Quantity": qty,
            "Unit Price": unit_price,
            "Total Amount": total,
            "Region": random.choice(regions),
            "Sales Rep": random.choice(reps),
            "Status": random.choice(statuses),
        })

    df = pd.DataFrame(rows)

    # ── Inject dirty data ──────────────────────────────────────────────

    # Duplicate 12 rows
    dup_idx = random.sample(range(len(df)), 12)
    df = pd.concat([df, df.iloc[dup_idx]], ignore_index=True)

    # Missing Customer Email (10 rows)
    miss_email_idx = random.sample(range(len(df)), 10)
    df.loc[miss_email_idx, "Customer Email"] = np.nan

    # Invalid email format (8 rows) — only where email is not already NaN
    valid_email_idx = df[df["Customer Email"].notna()].index.tolist()
    bad_email_idx = random.sample(valid_email_idx, 8)
    df.loc[bad_email_idx, "Customer Email"] = df.loc[bad_email_idx, "Customer Email"].apply(corrupt_email)

    # Missing Quantity (7 rows)
    df.loc[random.sample(range(len(df)), 7), "Quantity"] = np.nan

    # Non-numeric in Total Amount (5 rows) — must cast column to object first
    df["Total Amount"] = df["Total Amount"].astype(object)
    df.loc[random.sample(range(len(df)), 5), "Total Amount"] = "N/A"

    # Extra whitespace in Customer Name (15 rows)
    name_idx = random.sample(range(len(df)), 15)
    df.loc[name_idx, "Customer Name"] = "  " + df.loc[name_idx, "Customer Name"] + "  "

    # Fully empty row (3 rows)
    for _ in range(3):
        df = pd.concat([df, pd.DataFrame([{}])], ignore_index=True)

    df = df.sample(frac=1, random_state=SEED).reset_index(drop=True)
    return df


# ── Customer dataset ──────────────────────────────────────────────────────────

def generate_customers(n: int = 160) -> pd.DataFrame:
    companies = ["Apex Dynamics", "Bluewave Corp", "Crestline Partners", "Delta Systems",
                 "Emerald Solutions", "Frontier Tech", "Granite Holdings", "Harbor Industries",
                 "Ironclad Networks", "Juniper Group"]
    cities = ["New York", "Los Angeles", "Chicago", "Houston", "Phoenix",
              "Philadelphia", "San Antonio", "San Diego", "Dallas", "San Jose"]
    countries = ["USA", "Canada", "UK", "Australia", "Germany"]
    statuses = ["Active", "Active", "Active", "Inactive", "Prospect", "Churned"]
    first_names = ["Emma", "Liam", "Olivia", "Noah", "Ava", "Ethan", "Sophia", "Mason",
                   "Isabella", "Logan", "Mia", "Lucas", "Charlotte", "Aiden", "Amelia"]
    last_names = ["Roberts", "Clark", "Lewis", "Robinson", "Walker", "Hall", "Allen",
                  "Young", "Hernandez", "King", "Wright", "Scott", "Torres", "Nguyen"]
    domains = ["enterprise.com", "corp.org", "business.net", "holdings.io", "group.co"]

    rows = []
    for i in range(1, n + 1):
        fname = random.choice(first_names)
        lname = random.choice(last_names)
        company = random.choice(companies)
        email = f"{fname.lower()[0]}{lname.lower()}@{company.lower().replace(' ', '')}.{random.choice(['com','net','org'])}"
        rows.append({
            "Customer ID": f"CUS-{i:04d}",
            "Name": f"{fname} {lname}",
            "Email": email,
            "Phone": f"+1-{random.randint(200,999)}-{random.randint(200,999)}-{random.randint(1000,9999)}",
            "Company": company,
            "City": random.choice(cities),
            "Country": random.choice(countries),
            "Annual Revenue": round(random.uniform(10_000, 5_000_000), 2),
            "Join Date": rand_date(date(2018, 1, 1), date(2024, 6, 30)),
            "Account Status": random.choice(statuses),
        })

    df = pd.DataFrame(rows)

    # ── Inject dirty data ──────────────────────────────────────────────

    # Duplicate 8 rows
    dup_idx = random.sample(range(len(df)), 8)
    df = pd.concat([df, df.iloc[dup_idx]], ignore_index=True)

    # Missing Email (10 rows)
    df.loc[random.sample(range(len(df)), 10), "Email"] = np.nan

    # Invalid emails (6 rows)
    valid_idx = df[df["Email"].notna()].index.tolist()
    df.loc[random.sample(valid_idx, 6), "Email"] = df.loc[
        random.sample(valid_idx, 6), "Email"
    ].apply(corrupt_email)

    # Missing Annual Revenue (18 rows)
    df.loc[random.sample(range(len(df)), 18), "Annual Revenue"] = np.nan

    # Non-numeric in Annual Revenue (for numeric validation demo)
    df["Annual Revenue"] = df["Annual Revenue"].astype(object)
    df.loc[random.sample(range(len(df)), 4), "Annual Revenue"] = "undisclosed"

    # Extra whitespace in Name
    name_idx = random.sample(range(len(df)), 12)
    df.loc[name_idx, "Name"] = df.loc[name_idx, "Name"].str.upper()

    df = df.sample(frac=1, random_state=SEED).reset_index(drop=True)
    return df


# ── Main ──────────────────────────────────────────────────────────────────────

def write_styled_excel(df: pd.DataFrame, path: Path, sheet_name: str) -> None:
    """Write a DataFrame to a formatted .xlsx file with a styled header row."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)
        wb = writer.book
        ws = writer.sheets[sheet_name]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
        thin = Side(style="thin", color="BFBFBF")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        even_fill = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=1):
            fill = even_fill if i % 2 == 0 else PatternFill()
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
                cell.border = border

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        ws.freeze_panes = "A2"


def main():
    print("Generating sample datasets...")

    sales = generate_sales()

    # CSV version
    sales_csv = DATA_DIR / "sample_sales.csv"
    sales.to_csv(sales_csv, index=False)
    print(f"  ✓ {sales_csv.name}  ({len(sales)} rows)")

    # Excel version — primary format for an Excel Automation tool
    sales_xlsx = DATA_DIR / "sample_sales.xlsx"
    write_styled_excel(sales, sales_xlsx, "Sales Data")
    print(f"  ✓ {sales_xlsx.name}  ({len(sales)} rows, formatted)")

    customers = generate_customers()

    # CSV version
    cust_csv = DATA_DIR / "sample_customers.csv"
    customers.to_csv(cust_csv, index=False)
    print(f"  ✓ {cust_csv.name}  ({len(customers)} rows)")

    # Excel version
    cust_xlsx = DATA_DIR / "sample_customers.xlsx"
    write_styled_excel(customers, cust_xlsx, "Customer Data")
    print(f"  ✓ {cust_xlsx.name}  ({len(customers)} rows, formatted)")

    print("Done. Sample files are ready in /data/")


if __name__ == "__main__":
    main()
