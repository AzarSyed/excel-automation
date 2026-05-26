"""Excel export module — generates formatted, multi-sheet workbooks."""
import io
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils.dataframe import dataframe_to_rows

from utils.logger import setup_logger

logger = setup_logger("excel_exporter")

# ── Shared style constants ────────────────────────────────────────────────────
_HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
_TITLE_FONT = Font(bold=True, size=14, color="1F3864", name="Calibri")
_SUBTITLE_FONT = Font(italic=True, size=9, color="888888", name="Calibri")
_EVEN_FILL = PatternFill(start_color="D9E8F5", end_color="D9E8F5", fill_type="solid")
_THIN = Side(style="thin", color="BFBFBF")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALERT_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")


class ExcelExporter:
    """Builds production-style Excel workbooks with multiple report sheets."""

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> None:
        """Write a DataFrame to a worksheet starting at start_row."""
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

    @staticmethod
    def _auto_width(ws) -> None:
        """Set column widths based on content (capped at 60)."""
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value or "")) for cell in col), default=10
            )
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    def _style_table(self, ws, header_row: int = 1) -> None:
        """Apply header style and alternating row fills to a data table."""
        for cell in ws[header_row]:
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = _BORDER

        for i, row in enumerate(
            ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row), start=1
        ):
            fill = _EVEN_FILL if i % 2 == 0 else PatternFill()
            for cell in row:
                cell.fill = fill
                cell.alignment = Alignment(vertical="center")
                cell.border = _BORDER

        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    def _add_title(self, ws, title: str, subtitle: str = "") -> int:
        """Insert a 2-row title block and return the next available row number."""
        ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
        ws.cell(row=2, column=1, value=subtitle or f"Generated: {datetime.now():%Y-%m-%d %H:%M}").font = _SUBTITLE_FONT
        ws.row_dimensions[1].height = 22
        return 4  # data starts at row 4

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def _sheet_cleaned(self, wb, df: pd.DataFrame) -> None:
        ws = wb.active
        ws.title = "Cleaned Data"
        data_row = self._add_title(ws, "Cleaned Dataset", f"{len(df):,} rows after cleaning")
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=data_row):
            for col_idx, value in enumerate(row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        self._style_table(ws, header_row=data_row)
        self._auto_width(ws)

    def _sheet_invalid(self, wb, df: pd.DataFrame) -> None:
        ws = wb.create_sheet("Invalid Records")
        if df.empty:
            ws["A1"] = "No invalid records — all rows passed validation."
            ws["A1"].font = Font(bold=True, color="27AE60")
            return
        data_row = self._add_title(ws, "Invalid Records", f"{len(df):,} rows flagged")
        for row_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=data_row):
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx > data_row:
                    cell.fill = _ALERT_FILL
        self._style_table(ws, header_row=data_row)
        self._auto_width(ws)

    def _sheet_analytics(self, wb, analytics: dict) -> None:
        ws = wb.create_sheet("Analytics Summary")
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 32
        ws.column_dimensions["C"].width = 22

        self._add_title(ws, "Analytics Summary")

        row = 4
        for key, value in analytics.items():
            label = str(key).replace("_", " ").title()
            label_cell = ws.cell(row=row, column=1, value=label)
            label_cell.font = Font(bold=True, color="1F3864")
            label_cell.border = _BORDER

            if isinstance(value, dict):
                for sub_k, sub_v in value.items():
                    ws.cell(row=row, column=2, value=str(sub_k)).border = _BORDER
                    ws.cell(row=row, column=3, value=str(sub_v)).border = _BORDER
                    row += 1
            else:
                ws.cell(row=row, column=2, value=str(value)).border = _BORDER
                row += 1

        # Simple bar chart for missing values
        missing = analytics.get("missing_by_column", {})
        missing_nonzero = {k: v for k, v in missing.items() if v > 0}
        if missing_nonzero:
            chart_ws = wb.create_sheet("Missing Values Chart")
            chart_ws["A1"] = "Column"
            chart_ws["B1"] = "Missing Count"
            for i, (col, cnt) in enumerate(missing_nonzero.items(), start=2):
                chart_ws.cell(row=i, column=1, value=col)
                chart_ws.cell(row=i, column=2, value=cnt)

            chart = BarChart()
            chart.title = "Missing Values by Column"
            chart.y_axis.title = "Count"
            chart.x_axis.title = "Column"
            chart.style = 10
            chart.shape = 4
            n = len(missing_nonzero) + 1
            data_ref = Reference(chart_ws, min_col=2, min_row=1, max_row=n)
            cats = Reference(chart_ws, min_col=1, min_row=2, max_row=n)
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats)
            chart_ws.add_chart(chart, "D2")

    def _sheet_log(self, wb, cleaning_log: list[str]) -> None:
        ws = wb.create_sheet("Processing Log")
        ws.column_dimensions["A"].width = 90
        self._add_title(ws, "Processing Log")
        for i, entry in enumerate(cleaning_log, start=4):
            cell = ws.cell(row=i, column=1, value=f"• {entry}")
            cell.alignment = Alignment(wrap_text=True)
            if i % 2 == 0:
                cell.fill = _EVEN_FILL

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def export_full_report(
        self,
        cleaned_df: pd.DataFrame,
        invalid_df: pd.DataFrame,
        analytics: dict,
        cleaning_log: list[str],
    ) -> bytes:
        """
        Build a formatted, multi-sheet Excel workbook and return it as bytes.

        Sheets:
          1. Cleaned Data
          2. Invalid Records
          3. Analytics Summary  (+  Missing Values Chart if applicable)
          4. Processing Log
        """
        wb = openpyxl.Workbook()

        self._sheet_cleaned(wb, cleaned_df)
        self._sheet_invalid(wb, invalid_df)
        self._sheet_analytics(wb, analytics)
        self._sheet_log(wb, cleaning_log)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        logger.info(
            "Excel report exported — %d clean rows, %d invalid rows",
            len(cleaned_df),
            len(invalid_df),
        )
        return buf.getvalue()
